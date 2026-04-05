import os
from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from openpyxl import load_workbook

API_TOKEN = os.getenv("BOT_TOKEN")

if not API_TOKEN:
    raise ValueError("❌ BOT_TOKEN не найден! Добавь его в Render")

bot = Bot(token=API_TOKEN)
dp = Dispatcher(bot)

STAGES = {
    1: "🛒 Заказ оформлен",
    2: "🇺🇸 На складе в США",
    3: "🇪🇺 Транзитный рейс в Европе",
    4: "🇷🇺 Таможня в Москве",
    5: "✅ Доставлен в РФ"
}

def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip()

    # убираем частый excel-формат типа 2.0
    if text.endswith(".0"):
        text = text[:-2]

    # убираем неразрывные пробелы и лишние пробелы
    text = text.replace("\xa0", "").strip()

    return text

def normalize_status(value):
    text = normalize_text(value)

    if not text:
        return None

    try:
        return int(float(text))
    except ValueError:
        return None

def find_order(order_id):
    wb = load_workbook("orders1.xlsx", data_only=True)
    ws = wb.active

    search_id = normalize_text(order_id)

    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        row_id = normalize_text(row[0])
        row_status = normalize_status(row[3])

        if row_id == search_id:
            return {
                "order_id": row_id,
                "status": row_status
            }

    return None

@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("📦 Отследить заказ")
    await message.answer("Выберите действие:", reply_markup=kb)

@dp.message_handler(lambda m: m.text == "📦 Отследить заказ")
async def track(message: types.Message):
    await message.answer("Введите номер заказа:")

@dp.message_handler()
async def handle(message: types.Message):
    order_id = message.text.strip()
    order = find_order(order_id)

    if not order:
        await message.answer("❌ Заказ не найден")
        return

    stage = order["status"]

    if stage not in STAGES:
        await message.answer("❌ У заказа некорректный статус")
        return

    current_status = STAGES[stage]

    text = f"""📦 Заказ: {order['order_id']}

📍 Статус:
{current_status}"""

    await message.answer(text)

if __name__ == "__main__":
    async def on_startup(dp):
        await bot.delete_webhook(drop_pending_updates=True)

    print("🚀 БОТ ЗАПУЩЕН")
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)

